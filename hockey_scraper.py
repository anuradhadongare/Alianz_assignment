##The main scraping and processing script to scrap from the website
import asyncio
import aiohttp
from bs4 import BeautifulSoup
import zipfile
import openpyxl
from collections import defaultdict

#####################Defines an asynchronous function fetch_page which is responsible for fetching the HTML content of a given URL#############################################
async def fetch_page(session, url, page_num):
    ###This uses an asynchronous context manager to make a GET request to the specified url using the provided session. 
    ###The async with ensures the response is properly closed, even if errors occur.
    async with session.get(url) as response:
        ##Checks if the HTTP status code is 200, which indicates a successful response.
        if response.status == 200:
            #If the request was successful, this line retrieves the HTML content from the response using await response.text(). 
            #The await keyword pauses execution until the response is fully received.
            html = await response.text()
            #This opens a zip file named "html_files.zip" in append mode ("a"). The with statement ensures the zip file is properly closed.
            with zipfile.ZipFile("html_files.zip", "a") as zf:
                # This writes the HTML content to the zip file as a new file named page_num.html.  
                zf.writestr(f"{page_num}.html", html)  # Store in zip
            return html ## the function returns the HTML content.
        else: #status code is not 200 (indicating an error).
            print(f"Error fetching {url}: {response.status}")
            return None # None to indicate that the page fetch was unsuccessful.
            
            
###################### Defines an asynchronous function named scrape_data for webscrapping ###############################################################
async def scrape_data(session, base_url, num_pages):
    # Initializes an empty list called all_data. This list will store the scraped data from all pages.
    all_data = []
    #this loop iterates through each page number, from 1 to num_pages
    for page in range(1, num_pages + 1):
        #Constructs the URL for the current page by appending the page number as a query parameter to the base_url. 
        url = f"{base_url}?page={page}"
        #This line calls asynchronous function called fetch_page to fetch the HTML content of the current page. 
        html = await fetch_page(session, url, page)
        # if html is None or empty the code inside this block is skipped.
        if html:
            #  BeautifulSoup is a library used for parsing HTML and XML.  "html.parser" specifies the HTML parser to use.
            soup = BeautifulSoup(html, "html.parser")
            # This line uses BeautifulSoup's find() method to search for a <table> element with the class attribute "table".  This assumes the data is in an HTML table.
            table = soup.find("table", class_="table")

            #If no table is found, the code inside this block is skipped.
            if table:
                #Finds all the table rows (<tr> elements) within the table. [1:] slices the list to exclude the first row, which is assumed to be the table header.
                rows = table.find_all("tr")[1:] # Skip header
                #This loop iterates through each row in the table (excluding the header)
                for row in rows:
                    #Finds all the table data cells (<td> elements) within the current row.
                    cells = row.find_all("td")
                    #Extracts the text content from each cell, strips any leading/trailing whitespace, and stores it in a list called row_data. 
                    #This list represents the data for a single row
                    row_data = [cell.text.strip() for cell in cells]
                    #Appends the row_data (the data for the current row) to the all_data list.
                    all_data.append(row_data)
        #This line is very important for ethical scraping. It pauses the execution for 0.1 seconds (100 milliseconds) between requests. 
        # Scraping too aggressively can get your IP address blocked.
        await asyncio.sleep(0.1)
    ##After processing all pages, the function returns the all_data list, which contains all the scraped data.
    return all_data

##Defines the analyze_data function, which processes the raw scraped data to find the team with the most wins (winner) and the team with the fewest wins (loser) for each year. 
def analyze_data(data):
    #Creates a dictionary where, if you try to access a key that doesn't exist, it automatically creates that key and assigns it the value returned by the lambda function
    #This lambda function creates a dictionary with two keys: "wins" and "losses". Each of these keys is also a defaultdict(int). 
    yearly_stats = defaultdict(lambda: {"wins": defaultdict(int), "losses": defaultdict(int)})

    #iterates through each row of the scraped data.
    for row in data:
        year = int(row[1].split('-')[0]) #  Extracts the year from the second element (row[1]) of the current row.  
        team = row[0]    # Extracts the team name from the first element (row[0]) of the current row.
        wins = int(row[2])  # Extracts the number of wins from the third element (row[2]) of the current row and converts it to an integer
        #This line updates the yearly_stats dictionary.  It adds the wins for the current team to the count for the current year.
        #Because yearly_stats is a nested defaultdict, if the year or team doesn't exist as a key, it will be automatically created.
        yearly_stats[year]["wins"][team] += wins

    summary = [] #empty list called summary.
    #this loop iterates through the yearly_stats dictionary, where year is the year and stats is the nested dictionary containing the "wins" and "losses" data for that year. 
    for year, stats in yearly_stats.items():
      wins_data = stats["wins"]  #Gets the "wins" data for the current year.
      if wins_data: # Check if data exists for the year
        winner = max(wins_data, key=wins_data.get)  # Finds the team with the maximum number of wins using the max()
        winner_wins = wins_data[winner]  # Gets the number of wins for the winner.
        loser = min(wins_data, key=wins_data.get) #Finds the team with the minimum number of wins using the min()
        loser_wins = wins_data[loser]  # Gets the number of wins for the loser.          
        summary.append([year, winner, winner_wins, loser, loser_wins])
          
    #summary list, which contains the summarized data for all years.
    return summary

# async with : This is an asynchronous context manager. It ensures that the session is properly closed even if errors occur. 
##Asynchronous operations allow your program to perform other tasks while waiting for network requests to complete, making the scraping process much faster.

async def main():
    # This line defines the base URL of the website you want to scrape.
    base_url = "https://www.scrapethissite.com/pages/forms/"
    num_pages = 24  # Determined by inspecting the site as 24 pages
    
    #This line initiates an asynchronous HTTP client session using aiohttp and creates aobject
    async with aiohttp.ClientSession() as session:
        #responsible for actually fetching and parsing the data from the website and storing in the object all_data
        all_data = await scrape_data(session, base_url, num_pages)

    if all_data: # if not empty
        ##: This creates a new Excel workbook object using the openpyxl library.  This is the starting point for creating the Excel file.
        workbook = openpyxl.Workbook()
        #a new workbook has one sheet, and this line gets a reference to it.
        sheet1 = workbook.active 
        #This sets the title of the worksheet to "NHL Stats 1990-2011". 
        sheet1.title = "NHL Stats 1990-2011"        
        sheet1.append(["Team", "Season", "Wins", "Losses"]) # adds a header row to the worksheet

        ##This loop iterates through each item in the all_data list. 
        #It's assumed that all_data is a list of lists (or a list of tuples), where each inner list represents a row of data scraped from the website
        for row in all_data:
            sheet1.append(row)
        #This line calls a function named analyze_data (which is not defined in this snippet) and passes the all_data (the scraped data) as an argument.
        summary_data = analyze_data(all_data)
        ##This creates a new worksheet in the Excel workbook and names it "Winner and Loser per Year".  The create_sheet() method creates a new sheet and returns a sheet object.
        sheet2 = workbook.create_sheet("Winner and Loser per Year") ##
        sheet2.append(["Year", "Winner", "Winner Num. of Wins", "Loser", "Loser Num. of Wins"])   ##This adds a header row to the new worksheet.  The
        # This loop iterates through each item in the summary_data list.
        for row in summary_data:
            sheet2.append(row) ##this line appends each row from summary_data to the worksheet

        workbook.save("hockey_stats.xlsx") ## This calls the save() method of the openpyxl.Workbook object
        print("Data scraped and saved to hockey_stats.xlsx and html_files.zip")

if __name__ == "__main__":
    #asyncio.run primary way to run an asynchronous program in Python. main() function is where you would typically call the other asynchronous functions
    asyncio.run(main())  
